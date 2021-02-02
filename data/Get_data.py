# -*- coding: UTF-8 -*-

import os
import openpyxl
from configparser import ConfigParser

class getdata:
    def __init__(self):
        #获取config基础数据
        file = os.path.abspath(os.path.join(os.getcwd(),'..','database','config'))
        config = ConfigParser()
        config.read(file)
        self.CaseId_col = config.get('excel','Case_id')
        self.APIname_col = config.get('excel','API_name')
        self.Url_col = config.get('excel', 'Url')
        self.Method_col = config.get('excel', 'Method')
        self.RequestDataType_col = config.get('excel', 'Request_Data_Type')
        self.RequestData_col = config.get('excel', 'Request_Data')
        self.StatusCode_col = config.get('excel', 'Status_Code')
        self.token_col = config.get('excel','token')
        self.CheckPoint_col = config.get('excel', 'Check_Point')
        self.error_col = config.get('excel', 'error')
        #需要根据项目所在路径配置地址
        self.filename = os.path.join('D:\python-UI\jiekou\data','test.xlsx')
        # 打开工作簿
        self.table = openpyxl.load_workbook(self.filename)

        #1.获取表单方法1
        # self.workBook = self.table.active

        #2.获取表单方法2
        sheets = self.table.get_sheet_names()
        self.wk = sheets[0]
        self.workBook = self.table[self.wk]

    #获取总行数
    def get_case_lines(self):
        rowNum = self.workBook.max_row
        return rowNum

    #获取caseId对应的行所有值
    def get_row_data(self,case_id):
        row_list = []
        for i in self.workBook[case_id]:
            row_list.append(i.value)
        return row_list

    #获取总列数
    def get_case_col(self):
        colNum = self.workBook.max_column
        return colNum

    #获取某个单元格数据
    def get_cell_value(self,row,col):
        cellValue = self.workBook.cell(row,col).value
        return cellValue

    # 获取单元格Case_id
    def get_Case_id(self, row):
        col = int(self.CaseId_col)
        Case_id = self.get_cell_value(row,col)
        return Case_id

    # 获取单元格API_name
    def get_API_name(self, row):
        col = int(self.APIname_col)
        Case_id = self.get_cell_value(row,col)
        return Case_id

    #获取单元格URL
    def get_Url(self,row):
        col = int(self.Url_col)
        Url = self.get_cell_value(row,col)
        return Url

    # 获取单元格请求方式
    def get_Method(self,row):
        col = int(self.Method_col)
        Method = self.get_cell_value(row,col)
        return Method

    # 获取单元格请求格式
    def get_Request_Data_Type(self,row):
        col = int(self.RequestDataType_col)
        Request_Data_Type= self.get_cell_value(row,col)
        return Request_Data_Type

    # 获取单元格请求参数
    def get_Request_Data(self,row):
        col = int(self.RequestData_col)
        Request_Data = self.get_cell_value(row,col)
        if Request_Data == ''or Request_Data == None:
            return None
        else:
            return Request_Data

    # 获取请求状态
    def get_Status_Code(self, row):
        col = int(self.StatusCode_col)
        Status_Code = self.get_cell_value(row,col)
        return Status_Code


    # 获取请求token
    def get_token(self, row):
        col = int(self.token_col)
        token = self.get_cell_value(row, col)
        return token

    # 获取请求检查点
    def get_Check_Point(self, row):
        col = int(self.CheckPoint_col)
        Check_Point = self.get_cell_value(row,col)
        return Check_Point


    # 获取请求error
    def get_error(self, row):
        col = int(self.error_col)
        error = self.get_cell_value(row,col)
        return error




if __name__ == '__main__':
    getdata()
    # t = getdata().get_case_lines()
    # print(t)
#     getdata().get_Url(2)
#     t = getdata().get_Request_Data_Type(2)
#     print(t)
    # getdata().get_Method(2)
    # getdata().get_Case_id(3)
    # a= getdata().get_cell_value(row=2,col=1)
    # getdata().get_case_col(2)
    # getdata().get_row_data(1)
    # getdata().get_row_data(3)
